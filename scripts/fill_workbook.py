"""form1_proposal.xlsx（様式1＿研究計画調書.xlsx）への書き込みヘルパー

既存のテンプレートをコピーし、ヒアリング結果の dict を受けて対応セルに値を書く。
既存の数式・書式は保持する。openpyxl のみ使用。

使い方:
    python fill_workbook.py \\
        --template "<SKILL_ROOT>/call_materials/application_forms/form1_proposal.xlsx" \\
        --data data.json \\
        --output "<WORK_DIR>/様式１＿研究計画調書＿12345＿YAMADA_Taro.xlsx"

data.json の形式は build_payload() コメント参照。
"""

from __future__ import annotations
import argparse
import json
import shutil
from pathlib import Path
from openpyxl import load_workbook

# スキルルートは本ファイルの親ディレクトリの親（.../scripts/fill_workbook.py → .../）
SKILL_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_TEMPLATE = SKILL_ROOT / "call_materials" / "application_forms" / "form1_proposal.xlsx"

# ======== 1枚目のセル割当 ========
# （ラベルは A 列、入力値は B 列の結合セル開始点に書き込む）
SHEET1_CELLS = {
    # 基本情報
    "apply_date": ("研究計画調書_1枚目", "A5"),  # 令和8年XX月XX日
    "erad_researcher_number": ("研究計画調書_1枚目", "B6"),  # 8桁
    "email": ("研究計画調書_1枚目", "B7"),
    "name_kana": ("研究計画調書_1枚目", "D8"),  # フリガナ値 (B8="フリガナ" ラベル、D8:I8 が入力)
    "name_kanji": ("研究計画調書_1枚目", "D9"),
    "birthdate": ("研究計画調書_1枚目", "B10"),
    "erad_institution_code": ("研究計画調書_1枚目", "B11"),
    "institution": ("研究計画調書_1枚目", "B12"),
    "department": ("研究計画調書_1枚目", "B13"),
    "position": ("研究計画調書_1枚目", "B14"),
    "category": ("研究計画調書_1枚目", "B15"),  # 区分（リスト選択）
    "research_field": ("研究計画調書_1枚目", "B17"),
    "main_usecase": ("研究計画調書_1枚目", "B18"),
    "main_usecase_other": ("研究計画調書_1枚目", "B19"),
    # 研究課題名
    "title_ja": ("研究計画調書_1枚目", "B22"),
    "title_en": ("研究計画調書_1枚目", "B22"),  # B22 merged B22:K22; 日本語のみが入る設計
    # 現在の具体的な活用方法
    "ai_usage_ja": ("研究計画調書_1枚目", "B26"),
}

# サブユースケース（B20,D20,F20,H20,J20, B21,D21,F21,H21）は順番に 8 項目
SUBCASE_CELLS = [
    ("B20", "1.学習用データセット構築"),
    ("D20", "2.既存モデルの適応"),
    ("F20", "3.AIモデル開発"),
    ("H20", "4.既存モデル評価"),
    ("J20", "5.実験自動化・自律化"),
    ("B21", "6.シミュレーション・デジタルツイン"),
    ("D21", "7.発見・設計支援"),
    ("F21", "8.高度データ解析・モデリング"),
]

# 現在のAI利活用度合い（10項目）
AI_USAGE_CELLS = [
    ("B24", "研究でAIをまったく使っていない"),
    ("D24", "研究そのもの以外の業務でAIを使っている"),
    ("F24", "文献探索や要約にAIを使っている"),
    ("H24", "論文執筆や発表資料にAIを使っている"),
    ("J24", "仮説検討やアイデア出しにAIを使っている"),
    ("B25", "自作コード含めAIでデータ分析"),
    ("D25", "AI分析結果を論文・発表で発表"),
    ("F25", "APIで既存AIを研究プロセスに組み込み"),
    ("H25", "AIモデルの開発経験あり"),
    ("J25", "新しい基盤モデル開発の経験あり"),
]

# ======== 2枚目 ========
SHEET2_CELLS = {
    "purpose_ja": ("研究計画調書_2枚目", "A3"),
    "method_ja": ("研究計画調書_2枚目", "A5"),
    "ai_rationale_ja": ("研究計画調書_2枚目", "A7"),
    "goals_ja": ("研究計画調書_2枚目", "A9"),
    "knowhow_ja": ("研究計画調書_2枚目", "A11"),
    "achievements": ("研究計画調書_2枚目", "A15"),
}

# ======== 3枚目（予算明細） ========
# 各費目のヘッダー行（「事項」「品名・仕様」等のラベル行）を定義し、
# データ書き込みはヘッダー行 +1 から開始する。
# ハードコーディングを避け、ヘッダー行の定義を1箇所に集約する。
SHEET3_HEADER_ROWS = {
    "equipment": 5,       # 行5: 品名・仕様 / 設置機関 / 数量 / 単価 / 金額
    "consumables": 5,     # 行5: 事項 / 金額（設備備品費と同じヘッダー行）
    "honorarium": 14,     # 行14: 事項 / 金額
    "domestic_travel": 14,  # 行14: 事項 / 金額
    "foreign_travel": 14,   # 行14: 事項 / 金額
    "other": 22,          # 行22: 事項 / 金額
}

# 各費目のデータ開始行（ヘッダー行 + 1）
SHEET3_DATA_START = {k: v + 1 for k, v in SHEET3_HEADER_ROWS.items()}

# 各費目の最大データ行数
SHEET3_MAX_ROWS = {
    "equipment": 3,       # 行6〜8
    "consumables": 3,     # 行6〜8
    "honorarium": 2,      # 行15〜16
    "domestic_travel": 2, # 行15〜16
    "foreign_travel": 2,  # 行15〜16
    "other": 4,           # 行23〜26
}

SHEET3_TOTALS = {
    "equipment_total": ("研究計画調書_3枚目", "E9"),
    "consumables_total": ("研究計画調書_3枚目", "G9"),
    "honorarium_total": ("研究計画調書_3枚目", "B17"),
    "domestic_travel_total": ("研究計画調書_3枚目", "D17"),
    "foreign_travel_total": ("研究計画調書_3枚目", "F17"),
    "other_total": ("研究計画調書_3枚目", "B27"),
}

SHEET3_NECESSITIES = {
    "equipment_consumables_necessity": ("研究計画調書_3枚目", "A11"),  # merged A11:G11
    "honorarium_travel_necessity": ("研究計画調書_3枚目", "A19"),
    "other_necessity": ("研究計画調書_3枚目", "A29"),
}

# ======== 4枚目 ========
# API費用テーブル: 行5〜14、列 B(処理対象) C(金額千円) D(算定根拠、merged D:E)
# 計算資源費用テーブル: 行18〜27、列 B(GPU種類) C(選定理由) D(金額千円) E(算定根拠、merged D:E)
# ただし見た目では D14:E14 等が merged されている → 算定根拠は D列に書けばOK
SHEET4_API_START_ROW = 5
SHEET4_COMPUTE_START_ROW = 18


def build_empty_payload() -> dict:
    """ヒアリングで埋めていく辞書のひな形"""
    return {
        "filename": "様式１＿研究計画調書＿XXXXX＿YAMADA Taro.xlsx",
        # 1枚目
        "apply_date": "令和8年XX月XX日",
        "erad_researcher_number": "",
        "email": "",
        "name_kana": "",
        "name_kanji": "",
        "birthdate": "",
        "erad_institution_code": "",
        "institution": "",
        "department": "",
        "position": "",
        "category": "",
        "research_field": "",
        "main_usecase": "",
        "main_usecase_other": "",
        "subcases": [],   # 例: ["3.AIモデル開発", "8.高度データ解析・モデリング"]
        "title_ja": "",
        "title_en": "",
        "ai_usage_levels": [],  # 例: ["APIで既存AIを研究プロセスに組み込み"]
        "ai_usage_ja": "",
        "ai_usage_en": "",
        # 2枚目
        "purpose_ja": "",
        "purpose_en": "",
        "method_ja": "",
        "method_en": "",
        "ai_rationale_ja": "",
        "ai_rationale_en": "",
        "goals_ja": "",
        "goals_en": "",
        "knowhow_ja": "",
        "knowhow_en": "",
        "achievements": "",  # 改行区切り（最大5件）
        # 3枚目
        "equipment_rows": [],  # [{"name":..., "org":..., "qty":.., "unit_price":..,"amount":..}, ...]
        "consumables_rows": [],  # [{"item":..., "amount":..}, ...]
        "honorarium_rows": [],
        "domestic_travel_rows": [],
        "foreign_travel_rows": [],
        "other_rows": [],
        "equipment_consumables_necessity": "",
        "honorarium_travel_necessity": "",
        "other_necessity": "",
        # 4枚目
        "api_rows": [],  # [{"target":..., "amount":..., "basis":...}, ...]
        "compute_rows": [],  # [{"gpu":..., "rationale":..., "amount":..., "basis":...}, ...]
    }


def fill(template_path: str, data: dict, output_path: str) -> list:
    """テンプレをコピーしてデータを書き込む。警告・注意メッセージのリストを返す

    `rich_text=True` でロードすることで、セル内の部分的な赤字などの
    リッチテキスト装飾を保持したまま保存できる（これをしないと
    「(日本語：80文字以上400文字以内...)」などの赤字部分が黒字化する）。
    """
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path, rich_text=True)
    warnings = []

    def setv(sheet, coord, value):
        if value is None or value == "":
            return
        ws = wb[sheet]
        try:
            ws[coord] = value
        except Exception as e:
            warnings.append(f"セル書き込み失敗 {sheet}!{coord}: {e}")

    # ---- 1枚目 ----
    for key, (sheet, coord) in SHEET1_CELLS.items():
        if key == "title_en":
            continue  # B22 は日本語のみ保持（英語は別セルが無いため日本語に含めない方針）
        if key in data:
            setv(sheet, coord, data[key])

    # サブユースケース：True/False を順に書く
    subs = set(data.get("subcases", []))
    for coord, label in SUBCASE_CELLS:
        is_selected = any(label in s or s in label for s in subs)
        setv("研究計画調書_1枚目", coord, is_selected)

    # AI利活用度合い
    levels = set(data.get("ai_usage_levels", []))
    for coord, label in AI_USAGE_CELLS:
        is_selected = any(label[:10] in l or l[:10] in label for l in levels)
        setv("研究計画調書_1枚目", coord, is_selected)

    # ---- 2枚目 ----
    for key, (sheet, coord) in SHEET2_CELLS.items():
        if key in data:
            setv(sheet, coord, data[key])

    # ---- 3枚目 ----
    S3 = "研究計画調書_3枚目"

    # 設備備品費
    start = SHEET3_DATA_START["equipment"]
    max_count = SHEET3_MAX_ROWS["equipment"]
    for i, row in enumerate(data.get("equipment_rows", [])):
        r = start + i
        if i >= max_count:
            warnings.append(f"設備備品費が{max_count}行を超えています（行 insert が必要）: {row}")
            continue
        setv(S3, f"A{r}", row.get("name", ""))
        setv(S3, f"B{r}", row.get("org", ""))
        setv(S3, f"C{r}", row.get("qty"))
        setv(S3, f"D{r}", row.get("unit_price"))
        setv(S3, f"E{r}", row.get("amount"))

    start = SHEET3_DATA_START["consumables"]
    max_count = SHEET3_MAX_ROWS["consumables"]
    for i, row in enumerate(data.get("consumables_rows", [])):
        r = start + i
        if i >= max_count:
            warnings.append(f"消耗品費が{max_count}行超: {row}")
            continue
        setv(S3, f"F{r}", row.get("item", ""))
        setv(S3, f"G{r}", row.get("amount"))

    start = SHEET3_DATA_START["honorarium"]
    max_count = SHEET3_MAX_ROWS["honorarium"]
    for i, row in enumerate(data.get("honorarium_rows", [])):
        r = start + i
        if i >= max_count:
            warnings.append(f"謝金が{max_count}行超: {row}")
            continue
        setv(S3, f"A{r}", row.get("item", ""))
        setv(S3, f"B{r}", row.get("amount"))

    start = SHEET3_DATA_START["domestic_travel"]
    max_count = SHEET3_MAX_ROWS["domestic_travel"]
    for i, row in enumerate(data.get("domestic_travel_rows", [])):
        r = start + i
        if i >= max_count:
            warnings.append(f"国内旅費が{max_count}行超: {row}")
            continue
        setv(S3, f"C{r}", row.get("item", ""))
        setv(S3, f"D{r}", row.get("amount"))

    start = SHEET3_DATA_START["foreign_travel"]
    max_count = SHEET3_MAX_ROWS["foreign_travel"]
    for i, row in enumerate(data.get("foreign_travel_rows", [])):
        r = start + i
        if i >= max_count:
            warnings.append(f"外国旅費が{max_count}行超: {row}")
            continue
        setv(S3, f"E{r}", row.get("item", ""))
        setv(S3, f"F{r}", row.get("amount"))

    start = SHEET3_DATA_START["other"]
    max_count = SHEET3_MAX_ROWS["other"]
    for i, row in enumerate(data.get("other_rows", [])):
        r = start + i
        if i >= max_count:
            warnings.append(f"その他費目が{max_count}行超: {row}")
            continue
        setv(S3, f"A{r}", row.get("item", ""))
        setv(S3, f"B{r}", row.get("amount"))

    # 各費目の総計：数式ではなく数値で直接書き込む
    # （LibreOffice の recalc を通すとフォント色や条件付き書式が劣化するため、
    #   Python 側で合計を計算して数値で書き、recalc 不要にする）
    def sum_amount(rows, key="amount"):
        return sum((r.get(key) or 0) for r in rows)

    setv(S3, "E9", sum_amount(data.get("equipment_rows", [])))
    setv(S3, "G9", sum_amount(data.get("consumables_rows", [])))
    setv(S3, "B17", sum_amount(data.get("honorarium_rows", [])))
    setv(S3, "D17", sum_amount(data.get("domestic_travel_rows", [])))
    setv(S3, "F17", sum_amount(data.get("foreign_travel_rows", [])))
    setv(S3, "B27", sum_amount(data.get("other_rows", [])))

    # 必要性
    for key, (sheet, coord) in SHEET3_NECESSITIES.items():
        if key in data:
            setv(sheet, coord, data[key])

    # ---- 4枚目 ----
    for i, row in enumerate(data.get("api_rows", [])):
        r = SHEET4_API_START_ROW + i
        if r > 14:
            warnings.append(f"API費用が10行超: {row}")
            continue
        setv("研究計画調書_4枚目", f"B{r}", row.get("target", ""))
        setv("研究計画調書_4枚目", f"C{r}", row.get("amount"))
        setv("研究計画調書_4枚目", f"D{r}", row.get("basis", ""))

    for i, row in enumerate(data.get("compute_rows", [])):
        r = SHEET4_COMPUTE_START_ROW + i
        if r > 27:
            warnings.append(f"計算資源費用が10行超: {row}")
            continue
        setv("研究計画調書_4枚目", f"B{r}", row.get("gpu", ""))
        setv("研究計画調書_4枚目", f"C{r}", row.get("rationale", ""))
        setv("研究計画調書_4枚目", f"D{r}", row.get("amount"))
        setv("研究計画調書_4枚目", f"E{r}", row.get("basis", ""))

    wb.save(output_path)
    return warnings


def main():
    p = argparse.ArgumentParser()
    p.add_argument(
        "--template",
        default=str(DEFAULT_TEMPLATE),
        help=f"テンプレートパス（省略時はスキル同梱: {DEFAULT_TEMPLATE}）",
    )
    p.add_argument("--data", required=True, help="JSON file with build_empty_payload() schema")
    p.add_argument("--output", required=True)
    args = p.parse_args()

    with open(args.data, "r", encoding="utf-8") as f:
        data = json.load(f)

    warnings = fill(args.template, data, args.output)
    print(f"Wrote: {args.output}")
    if warnings:
        print("\n警告:")
        for w in warnings:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
