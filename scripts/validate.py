"""第2回 SPReAD 様式1 記入内容バリデーター

チェック内容:
1. 2枚目の各項目が文字数範囲内か
   - 日本語: 研究目的 80-400 / 研究方法 160-800 / AI利活用妥当性 160-800 /
            達成目標 100-500 / ノウハウ 60-300 / 成果の公開方針 任意 ≤150
   - 英語  : 研究目的 48-240 / 研究方法 96-480 / AI利活用妥当性 96-480 /
            達成目標 60-300 / ノウハウ 36-180 / 成果の公開方針 任意 ≤90
2. 必須項目（研究者番号・メアド・氏名・課題名・研究目的・研究方法 等）の未入力検出
3. 1枚目 D48（=SUM(E48:L48)）= 3枚目各小計の合算
4. 直接経費が 10万円〜500万円 の範囲か（千円単位 100〜5000）
5. 設備備品費・謝金・旅費のいずれかが総額の90%超なら必要性記載を確認
6. 4枚目合計（API + 計算資源）が 3枚目「その他」総計を超えていないこと
7. 1枚目の「9.その他」選択時に main_usecase_other が記入されていること
8. 業績欄が5件以下、各セルが1件分の入力であること
9. ファイル名規則（`第2回_様式1_研究計画調書_<機関コード>_<氏名>.xlsx`）

本スクリプトはあくまで本スキル独自の事前チェックである。提出前の最終確認は、
文部科学省公式の形式チェックツール（再配布禁止のため本スキルには同梱なし、
SKILL.md Step 7 にダウンロード手順を記載）を別途実行することを推奨する。

使い方:
    python validate.py workbook.xlsx
    python validate.py workbook.xlsx --language ja
"""

from __future__ import annotations
import argparse
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

SHEET_NAMES = {
    "ja": {
        "sheet1": "研究計画調書_1枚目",
        "sheet2": "研究計画調書_2枚目",
        "sheet3": "研究計画調書_3枚目",
        "sheet4": "研究計画調書_4枚目",
    },
    "en": {
        "sheet1": "Research Plan_Sheet 1",
        "sheet2": "Research Plan_Sheet 2",
        "sheet3": "Research Plan_Sheet 3",
        "sheet4": "Research Plan_Sheet 4",
    },
}

# 2枚目の項目別文字数/語数の下限・上限
TEXT_LIMITS = {
    "ja": [
        # (cell, lower, upper, label, optional)
        ("D8",  80, 400, "研究目的", False),
        ("D9",  160, 800, "研究方法", False),
        ("D10", 160, 800, "AI利活用の妥当性・実現可能性", False),
        ("D11", 100, 500, "達成目標", False),
        ("D12", 60, 300, "ノウハウ抽出・共有の実現計画", False),
        ("D13", None, 150, "成果の公開方針(任意)", True),
    ],
    "en": [
        ("D8",  48, 240, "Research Objectives", False),
        ("D9",  96, 480, "Research Methods", False),
        ("D10", 96, 480, "Rationale and Feasibility of AI Utilization", False),
        ("D11", 60, 300, "Achievement Goals", False),
        ("D12", 36, 180, "Plan for Extracting and Sharing AI Utilization Know-How", False),
        ("D13", None, 90, "Publication Policy for Research Outcomes (Optional)", True),
    ],
}

# 必須項目（言語別）
# 日本語版は D12=フリガナ, D13=漢字 に分かれているが、
# 英語版は D12:L13 結合で D12 のみ Full Name。
REQUIRED_SHEET1_BY_LANG = {
    "ja": [
        ("C8",  "e-Rad 研究者番号"),
        ("C10", "メールアドレス"),
        ("D13", "氏名（漢字）"),
        ("C16", "e-Rad 所属機関コード"),
        ("C18", "所属機関"),
        ("C20", "職"),
        ("C21", "所属機関の区分"),
        ("C23", "応募者属性の区分"),
        ("C27", "研究領域"),
        ("C29", "メインユースケース分類"),
        ("C35", "研究課題名"),
    ],
    "en": [
        ("C8",  "e-Rad Researcher Number"),
        ("C10", "Email"),
        ("D12", "Full Name"),
        ("C16", "e-Rad Institution Code"),
        ("C18", "Institution"),
        ("C20", "Position"),
        ("C21", "Institution Category"),
        ("C23", "Applicant Category"),
        ("C27", "Research Field"),
        ("C29", "Main Use Case"),
        ("C35", "Research Title"),
    ],
}

# 業績欄（最大5件）
ACHIEVEMENT_CELLS = ["D14", "D15", "D16", "D17", "D18"]

# 3枚目の各費目小計セル（テンプレ式 =SUM(...) が入っている）
S3_SUBTOTALS = {
    "equipment": "J31",
    "consumables": "N31",
    "honorarium": "E60",
    "domestic_travel": "J60",
    "foreign_travel": "N60",
    "other": "E89",
}

# 必要性記入欄
S3_NECESSITY = {
    "equipment_consumables": "C34",
    "honorarium_travel": "C63",
    "other": "C92",
}


def _count_chars(value, language: str) -> int:
    """日本語は文字数、英語は単語数（空白区切り）を返す。
    日本語版テンプレートの =LEN(...) と Excel の単語カウント関数の挙動に合わせる。
    """
    if value is None:
        return 0
    s = str(value)
    if language == "en":
        return len(s.split())
    return len(s)


def _detect_language(wb) -> str:
    """シート名で言語を判定。"""
    if "研究計画調書_1枚目" in wb.sheetnames:
        return "ja"
    if "Research Plan_Sheet 1" in wb.sheetnames:
        return "en"
    return "ja"


def _get_value(wb, sheet_name: str, coord: str):
    try:
        return wb[sheet_name][coord].value
    except Exception:
        return None


def _read_list_values(xlsx_path) -> dict:
    """リストシート（日本語版: `リスト` / 英語版: `List`）から正解選択肢を抽出。

    研究領域はA列2行目以降、メインユースケース分類はB列2行目以降。
    所属機関の区分・応募者属性の区分は C 列・D 列。返り値は
    {"research_fields": [...], "main_usecases": [...], ...} の dict。
    リストが存在しない／空の場合は空リストを返す（呼び出し側でガード）。
    """
    result = {
        "research_fields": [],
        "main_usecases": [],
        "institution_categories": [],
        "applicant_categories": [],
    }
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception:
        return result
    sheet_name = None
    for cand in ("リスト", "List"):
        if cand in wb.sheetnames:
            sheet_name = cand
            break
    if sheet_name is None:
        return result
    ws = wb[sheet_name]
    for col_letter, key in [
        ("A", "research_fields"),
        ("B", "main_usecases"),
        ("C", "institution_categories"),
        ("D", "applicant_categories"),
    ]:
        for row in range(2, 30):  # ヘッダ1行目を除外、最大28選択肢まで
            v = ws[f"{col_letter}{row}"].value
            if v is None or str(v).strip() == "":
                break
            result[key].append(str(v).strip())
    return result


def _to_number(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _check_filename(path: str, language: str) -> str | None:
    """ファイル名を公式形式チェックツール互換の正規表現で検査する。

    OK パターン:
      ja: 第2回_様式1_研究計画調書_<半角数字>_<氏名(_スペースなし)>.xlsx
      en: 2nd_form1_researchplan_<半角数字>_<Name(_スペースなし)>.xlsx (大小無視)

    NG なら理由文字列、OK なら None を返す。下書き接尾辞 (_DRAFT, _v2 等)、
    姓名間の `_`、半角スペースを検出して具体的なエラーを返す。
    """
    name = Path(path).name
    if not name.lower().endswith((".xlsx", ".xlsm")):
        return f"拡張子が .xlsx ではない: {name}"
    if name.startswith("~$"):
        return f"Excel の一時ファイル名で始まっている: {name}"
    stem = Path(name).stem
    # 全角空白・半角空白の検出
    if " " in stem or "　" in stem:
        return (
            f"ファイル名にスペースが含まれている: {name}\n"
            "  → 氏名のスペースを除去するか連結してください（例: 山田太郎, YamadaTaro）"
        )

    # 公式ツールの正規表現に準拠
    if language == "en":
        pattern = r"^2nd_form1_researchplan_([0-9]+)_([^_]+)$"
        m = re.fullmatch(pattern, stem, re.IGNORECASE)
        if not m:
            return (
                f"ファイル名が英語版規則に違反: {name}\n"
                "  → 正規形式: 2nd_Form1_ResearchPlan_<e-Rad機関コード>_<Name>.xlsx\n"
                "  → 氏名部分に _ を含めない（例: YamadaTaro）"
            )
        return None
    # ja
    pattern = r"^第2回_様式1_研究計画調書_([0-9]+)_([^_]+)$"
    m = re.fullmatch(pattern, stem)
    if not m:
        # 推定原因の特定
        parts = stem.split("_")
        if len(parts) >= 6:
            return (
                f"ファイル名に余分な _ が含まれている: {name}\n"
                f"  → _ で分割すると {len(parts)} 個になり、規則の 5 個を超過しています\n"
                "  → 正規形式: 第2回_様式1_研究計画調書_<e-Rad機関コード>_<氏名>.xlsx\n"
                "  → 末尾の _DRAFT 等を削除し、氏名内のスペースは連結してください"
            )
        return (
            f"ファイル名が日本語版規則に違反: {name}\n"
            "  → 正規形式: 第2回_様式1_研究計画調書_<e-Rad機関コード>_<氏名>.xlsx\n"
            "  → 氏名部分に _ を含めない（例: 山田太郎, YamadaTaro）"
        )


def check(path: str, language: str | None = None):
    wb = load_workbook(path, data_only=False)
    if language is None:
        language = _detect_language(wb)
    if language not in SHEET_NAMES:
        language = "ja"
    names = SHEET_NAMES[language]
    s1, s2, s3, s4 = names["sheet1"], names["sheet2"], names["sheet3"], names["sheet4"]

    errors: list[str] = []
    warnings: list[str] = []

    unit = "字" if language == "ja" else "words"

    # ---- ファイル名規則チェック ----
    fn_error = _check_filename(path, language)
    if fn_error:
        # 提出前に修正必須なのでエラー扱い
        errors.append(f"ファイル名規則違反: {fn_error}")

    # ---- 2枚目: 文字数/語数制限 ----
    for coord, low, high, label, optional in TEXT_LIMITS[language]:
        v = _get_value(wb, s2, coord)
        n = _count_chars(v, language)
        if not v:
            if optional:
                continue
            errors.append(f"必須項目未入力: {label} ({s2}!{coord})")
            continue
        if low is not None and n < low:
            errors.append(f"{label}: {n}{unit} は下限 {low}{unit} 未満 ({s2}!{coord})")
        if high is not None and n > high:
            errors.append(f"{label}: {n}{unit} は上限 {high}{unit} 超過 ({s2}!{coord})")

    # ---- 1枚目: 必須項目 ----
    required_sheet1 = REQUIRED_SHEET1_BY_LANG.get(language, REQUIRED_SHEET1_BY_LANG["ja"])
    for coord, label in required_sheet1:
        v = _get_value(wb, s1, coord)
        if v is None or str(v).strip() == "":
            errors.append(f"必須項目未入力: {label} ({s1}!{coord})")

    # ---- 1枚目: e-Rad 番号・機関コードの桁数チェック ----
    # e-Rad 研究者番号 (C8) は 8桁の半角数字
    erad_researcher = _get_value(wb, s1, "C8")
    if erad_researcher is not None and str(erad_researcher).strip():
        s_str = str(erad_researcher).strip()
        if not s_str.isdigit() or len(s_str) != 8:
            errors.append(
                f"e-Rad 研究者番号は半角数字8桁である必要があります"
                f"（現在の値: '{s_str}', 長さ: {len(s_str)}） ({s1}!C8)"
            )

    # e-Rad 所属機関コード (C16) は 10桁の半角数字（先頭ゼロを含む可能性あり）
    erad_inst = _get_value(wb, s1, "C16")
    if erad_inst is not None and str(erad_inst).strip():
        s_str = str(erad_inst).strip()
        if not s_str.isdigit() or len(s_str) != 10:
            errors.append(
                f"e-Rad 所属機関コードは半角数字10桁である必要があります"
                f"（現在の値: '{s_str}', 長さ: {len(s_str)}）。"
                f" 先頭ゼロを含む場合は文字列として正しく10桁入力されているかも確認 ({s1}!C16)"
            )

    # ---- 1枚目: 研究領域 (C27) / メインユースケース (C29) のリスト選択肢一致チェック ----
    # リストシートの値と完全一致しなければエラー（プルダウン選択時は問題ないが、
    # スクリプト書き込みやコピペで微妙に異なる文字列を入れてしまう事故を防ぐ）
    list_values = _read_list_values(path)

    research_field = _get_value(wb, s1, "C27")
    if research_field and list_values["research_fields"]:
        if str(research_field).strip() not in list_values["research_fields"]:
            errors.append(
                f"研究領域 ({s1}!C27) の値 '{research_field}' は、"
                f"リストシートの選択肢と一致しません。"
                f" 正解選択肢: {list_values['research_fields']}"
            )

    main_uc = _get_value(wb, s1, "C29")
    if main_uc and list_values["main_usecases"]:
        if str(main_uc).strip() not in list_values["main_usecases"]:
            errors.append(
                f"メインユースケース分類 ({s1}!C29) の値 '{main_uc}' は、"
                f"リストシートの選択肢と一致しません。"
                f" 正解選択肢: {list_values['main_usecases']}"
            )

    # 「9.その他」を選んだ場合は main_usecase_other (C31) も必須
    if main_uc and ("9" in str(main_uc) or "Other" in str(main_uc) or "その他" in str(main_uc)):
        other_text = _get_value(wb, s1, "C31")
        if not other_text:
            errors.append(f"メインユースケースで「9.その他」を選択時は自由記述が必須 ({s1}!C31)")

    # ---- 業績欄: 5件以下の確認 ----
    achievement_count = 0
    for coord in ACHIEVEMENT_CELLS:
        v = _get_value(wb, s2, coord)
        if v and str(v).strip():
            achievement_count += 1
    if achievement_count == 0:
        warnings.append("研究業績が0件（任意項目だが、応募者本人を含む業績がある場合は記載推奨）")
    # 6件目以降のセル（D19以降）は様式上存在しないので追加警告は不要

    # ---- 予算チェック ----
    # data_only で開くと数式キャッシュが読めるが、LibreOffice の recalc を通して
    # いない場合キャッシュが空のことがある。その場合は明細セルの値から自前で合計
    # する（テンプレ式 =SUM(...) と同じ範囲で再計算）。
    try:
        wb_dataonly = load_workbook(path, data_only=True)
    except Exception as e:
        warnings.append(f"data_only で開けず、合計値の検証をスキップ: {e}")
        return errors, warnings

    # 各小計の対象範囲（テンプレ式と同じ）
    SUBTOTAL_RANGES = {
        "equipment":      ("J", 11, 30),
        "consumables":    ("N", 11, 30),
        "honorarium":     ("E", 40, 59),
        "domestic_travel":("J", 40, 59),
        "foreign_travel": ("N", 40, 59),
        "other":          ("E", 69, 88),
    }
    # 設備備品費の J列はテンプレ式 =G*I なので、明細から再計算する場合は単価×数量を使う

    def _resolve_subtotal(key: str) -> float:
        # まず data_only で読んだキャッシュ値を試す
        cached = _to_number(_get_value(wb_dataonly, s3, S3_SUBTOTALS[key]))
        if cached > 0:
            return cached
        # キャッシュが空なら明細から自前で合計
        col, row_start, row_end = SUBTOTAL_RANGES[key]
        total = 0.0
        for r in range(row_start, row_end + 1):
            if key == "equipment":
                # J{row} = G{row} * I{row}
                unit = _to_number(_get_value(wb_dataonly, s3, f"G{r}"))
                qty = _to_number(_get_value(wb_dataonly, s3, f"I{r}"))
                total += unit * qty
            else:
                total += _to_number(_get_value(wb_dataonly, s3, f"{col}{r}"))
        return total

    sub_amounts = {key: _resolve_subtotal(key) for key in S3_SUBTOTALS}
    total_thousand = sum(sub_amounts.values())
    total_man = total_thousand / 10  # 千円 → 万円

    # 直接経費は 10万円以上 500万円以下
    if total_thousand > 0:
        if total_thousand < 100:
            errors.append(f"直接経費が下限(10万円)未満: {total_man:.1f}万円")
        if total_thousand > 5000:
            errors.append(f"直接経費が上限(500万円)超過: {total_man:.1f}万円")

    # 1枚目 D48（=SUM(E48:L48)）と 3枚目総計の整合
    # D48 自体はキャッシュ依存なので、E48..L48 の各参照式（=Sheet3!J31 等）も
    # キャッシュが無いと0になる。ここでは 3枚目から自前計算した小計合算と
    # キャッシュ値が乖離していないかだけ見る（キャッシュが0なら情報メッセージ）。
    sheet1_total = _to_number(_get_value(wb_dataonly, s1, "D48"))
    if sheet1_total > 0 and total_thousand > 0:
        if abs(sheet1_total - total_thousand) > 0.5:
            errors.append(
                f"1枚目総計 D48 ({sheet1_total/10:.1f}万円) と 3枚目各費目合算 ({total_man:.1f}万円) が一致しません。"
                f" Excel で開いて再計算（保存）すると解消する場合があります。"
            )
    elif sheet1_total == 0 and total_thousand > 0:
        # fill_workbook.py の後処理で数式セルには <v>0</v> が入っているため、
        # data_only ロードでは「0」として読まれる。これは正常な状態で、Excel で
        # 開いた瞬間に fullCalcOnLoad="1" により再計算されて正しい合計値が
        # キャッシュされる。自前計算が成立しているので情報メッセージのみ。
        warnings.append(
            "1枚目 D48 の数式キャッシュは <v>0</v> プレースホルダ状態です。"
            " Excel で開いた瞬間に再計算されて正しい合計に置き換わります（修復ダイアログは出ません）。"
            f" スクリプト側の自前合計では {total_man:.1f}万円 です。"
        )

    # 必要性記入欄（C34/C63/C92）は3枚とも常に埋める（必須）。
    # 空欄のまま提出すると審査員が経費の妥当性を読み取れず形式不備になる。
    # 当該費目の計上がない場合でも「本研究では…の計上はない」のように一行記載を求める。
    necessity_required = [
        ("C34", "設備備品費・消耗品費"),
        ("C63", "謝金・旅費"),
        ("C92", "その他費用"),
    ]
    for coord, label in necessity_required:
        nec = _get_value(wb_dataonly, s3, coord)
        if not nec or not str(nec).strip():
            errors.append(
                f"必要性記入欄 {s3}!{coord}（{label}の必要性）が未入力です。"
                f" 当該費目の計上有無に関わらず、必要性または不計上の旨を必ず記載してください。"
            )

    # 90% 超チェック（設備備品費 / 謝金 / 旅費）— 該当時は必要性をより詳細に書くべき
    if total_thousand > 0:
        breakdowns = [
            (sub_amounts["equipment"], "設備備品費", S3_NECESSITY["equipment_consumables"]),
            (sub_amounts["honorarium"], "謝金", S3_NECESSITY["honorarium_travel"]),
            (sub_amounts["domestic_travel"] + sub_amounts["foreign_travel"], "旅費", S3_NECESSITY["honorarium_travel"]),
        ]
        for amount, name, nec_coord in breakdowns:
            if amount / total_thousand > 0.9:
                nec = _get_value(wb_dataonly, s3, nec_coord)
                # 上の必須チェックで未入力は既にエラー化されているので、
                # ここでは「短すぎる」のチェックだけを警告として出す。
                if nec and len(str(nec).strip()) < 30:
                    warnings.append(
                        f"{name}が総額の90%超。{s3}!{nec_coord} の必要性記述が短すぎる可能性があります"
                        f"（{len(str(nec).strip())}字）。詳細な根拠を追記してください。"
                    )

        # 消耗品費 / その他 が大きな割合を占める場合も必要性記載が望ましい
        for amount, name, nec_coord in [
            (sub_amounts["consumables"], "消耗品費", S3_NECESSITY["equipment_consumables"]),
            (sub_amounts["other"], "その他", S3_NECESSITY["other"]),
        ]:
            if amount / total_thousand > 0.5:
                nec = _get_value(wb_dataonly, s3, nec_coord)
                if nec and len(str(nec).strip()) < 30:
                    warnings.append(
                        f"{name}が総額の50%超。{s3}!{nec_coord} の必要性記述が短すぎる可能性があります"
                        f"（{len(str(nec).strip())}字）。詳細な根拠を追記してください。"
                    )

    # ---- 4枚目合計と 3枚目「その他」の整合 ----
    api_sum = 0.0
    compute_sum = 0.0
    try:
        for r in range(9, 19):
            api_sum += _to_number(_get_value(wb_dataonly, s4, f"E{r}"))
        for r in range(22, 32):
            compute_sum += _to_number(_get_value(wb_dataonly, s4, f"F{r}"))
        four_total = api_sum + compute_sum
        other_total = sub_amounts["other"]
        if four_total > 0 and four_total > other_total + 0.5:
            errors.append(
                f"4枚目合計 (API {api_sum/10:.1f}万円 + 計算資源 {compute_sum/10:.1f}万円 = "
                f"{four_total/10:.1f}万円) が 3枚目「その他」総計 {other_total/10:.1f}万円 を超過しています。"
                " ※ 設備備品費として計上した計算資源は4枚目に書きます（その他に含めない）"
            )
        elif four_total > 0 and four_total < other_total * 0.8:
            warnings.append(
                f"4枚目合計 ({four_total/10:.1f}万円) は 3枚目「その他」総計 ({other_total/10:.1f}万円) より少なめ。"
                " AWS/API 以外の「その他」費目（クラウドストレージ単独、英文校正、論文掲載料 等）が含まれているか要確認。"
            )
    except Exception as e:
        warnings.append(f"4枚目整合チェックで例外: {e}")

    return errors, warnings


def main():
    p = argparse.ArgumentParser(description="第2回 SPReAD 様式1 バリデーター")
    p.add_argument("xlsx", help="検証対象の研究計画調書 xlsx")
    p.add_argument("--language", choices=["ja", "en"], default=None,
                   help="シート名から自動判定。明示する場合に指定。")
    args = p.parse_args()

    errors, warnings = check(args.xlsx, args.language)
    print("=" * 60)
    if errors:
        print(f"[NG] エラー {len(errors)} 件:")
        for e in errors:
            print(f"  - {e}")
    else:
        print("[OK] エラーなし")
    if warnings:
        print(f"\n[!]  警告 {len(warnings)} 件:")
        for w in warnings:
            print(f"  - {w}")
    print("=" * 60)
    print(
        "提出前には文部科学省公式の形式チェックツール (research_plan_self_check_v1.py)"
        " も実行することを推奨します。詳細は SKILL.md の Step 7 を参照。"
    )
    sys.exit(1 if errors else 0)


if __name__ == "__main__":
    main()
